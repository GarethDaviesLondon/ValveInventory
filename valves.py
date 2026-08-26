#!/usr/bin/env python3
"""
valves.py - valve stock inventory.

Command-line front end for the inventory: an argparse-based CLI with one
cmd_*(con, a) function per subcommand (box, find, show, search, add, edit,
take, move, bases, sock-add, sock-take, sock-move, set, merge, dupes, scan,
sheet, gaps, stats, export, import-csv). Each cmd_* function takes an open sqlite3
connection (`con`, row_factory=sqlite3.Row - see valvelib.init_db) and the
parsed argparse Namespace (`a`), and is responsible for its own commit().

This module shares its schema, DB connection helper, and the valve-type
classifier with the Tkinter desktop app (valves_gui.py) via valvelib.py -
both read and write the same valves.db, so avoid duplicating logic here that
belongs in valvelib.

  valves.py box 12                  what is in box 12
  valves.py find KT66               which boxes hold KT66 (follows equivalents)
  valves.py search --function pentode --heater 6.3 --pa '>20'
  valves.py add EL34 --box 1 --qty 4 --maker Svetlana --position B-12
  valves.py edit 417 --origin 'ex Bush DAC90' --test 'gm 9.8 mA/V'
  valves.py take EL34 --box 1 --qty 2
  valves.py show ECC83              full reference record
  valves.py set ECC83 --pa 1.2 --mu 100 --base B9A --confirm
  valves.py sheet ECC83             open the local datasheet
  valves.py scan                    link archive PDFs to types
  valves.py gaps                    types with no datasheet / no parameters
  valves.py export inventory.xlsx
"""

import argparse
import datetime
import os
import re
import subprocess
import sys

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import valvelib as V


# Per-lot detail fields (stock columns beyond box/qty/maker/condition/notes),
# as (column, CLI option, human label). One list so the add/edit options, the
# CSV importer, and the spreadsheet export stay in step with each other and
# with the schema.
LOT_FIELDS = [
    ("position", "position", "Position"),
    ("type1", "type1", "Type 1"),
    ("type2", "type2", "Type 2"),
    ("origin", "origin", "Origin"),
    ("test_values", "test", "Test values"),
    ("other", "other", "Other"),
]


# ---------------------------------------------------------------- helpers

def resolve(con, name):
    """Find matching type_key(s) from free-text user input.

    Tries three lookups in order, returning the first that finds anything:
    1. an exact match on the normalised key;
    2. a substring match on type_key (may return several types);
    3. a substring match against each type's equivalents list, so a search
       for e.g. '6L6' also finds types that merge() recorded 6L6 under.
    Returns an empty list if nothing matches at all.
    """
    key = V.norm(name)
    if not key:
        return []
    r = con.execute("SELECT type_key FROM valve_type WHERE type_key=?", (key,)).fetchone()
    if r:
        return [r["type_key"]]
    hits = [x["type_key"] for x in con.execute(
        "SELECT type_key FROM valve_type WHERE type_key LIKE ?", (f"%{key}%",))]
    if hits:
        return hits
    # search the equivalents lists
    return [x["type_key"] for x in con.execute(
        "SELECT type_key FROM valve_type WHERE UPPER(REPLACE(equivalents,' ','')) LIKE ?",
        (f"%{key}%",))]


def table(rows, cols):
    """Print `rows` (a list of dict-like records) as a plain-text table restricted to `cols`.

    Column widths are sized to the widest header/value in that column but
    capped at 46 characters; longer values are truncated with a trailing
    ellipsis so a single long field (e.g. notes) can't blow out the layout.
    """
    if not rows:
        print("  (nothing)")
        return
    w = {c: max(len(c), max(len(str(r[c] if r[c] is not None else "")) for r in rows)) for c in cols}
    w = {c: min(v, 46) for c, v in w.items()}
    print("  " + "  ".join(c.upper().ljust(w[c]) for c in cols))
    print("  " + "  ".join("-" * w[c] for c in cols))
    for r in rows:
        cells = []
        for c in cols:
            s = "" if r[c] is None else str(r[c])
            s = s if len(s) <= w[c] else s[:w[c] - 1] + "…"
            cells.append(s.ljust(w[c]))
        print("  " + "  ".join(cells))


def shown(rows, cols, always):
    """Drop columns from `cols` that are empty in every row of `rows`.

    The per-lot detail fields are optional and often unused, so listing them
    unconditionally would pad every table with blank columns for anyone who
    doesn't fill them in. Columns named in `always` are kept regardless, so
    the shape of a listing doesn't jump around as stock comes and goes.
    """
    return [c for c in cols
            if c in always or any(r.get(c) not in (None, "") for r in rows)]


def parse_cmp(expr):
    """'>20' -> ('>', 20.0);  '6.3' -> ('=', 6.3)"""
    m = re.match(r"^\s*(>=|<=|>|<|=)?\s*([\d.]+)\s*$", str(expr))
    if not m:
        raise ValueError(f"cannot parse comparison: {expr}")
    return (m.group(1) or "="), float(m.group(2))


# ---------------------------------------------------------------- commands

def cmd_box(con, a):
    """Print everything stored in one box: valve lots, sundry items, and bases/sockets.

    Lots are listed by position within the box (blank positions last, so a
    partly-positioned box still reads top to bottom) and carry their lot id,
    which is what 'edit' takes to change one.
    """
    rows = [dict(r) for r in con.execute(
        "SELECT id, position, type, type1, type2, qty, manufacturer, condition, "
        "origin, function, notes FROM v_stock WHERE box=? COLLATE NOCASE "
        "ORDER BY position IS NULL, position, type", (a.box,))]
    total = sum(r["qty"] for r in rows)
    print(f"\nBox {a.box} - {len(rows)} lots, {total} valves\n")
    table(rows, shown(rows, ["id", "position", "type", "type1", "type2", "qty",
                             "manufacturer", "condition", "origin", "function"],
                      always={"id", "type", "qty", "manufacturer", "condition", "function"}))
    sund = [dict(r) for r in con.execute(
        "SELECT description, qty FROM sundry WHERE box=? COLLATE NOCASE", (a.box,))]
    if sund:
        print("\n  other items:")
        table(sund, ["description", "qty"])
    bases = [dict(r) for r in con.execute(
        "SELECT base, qty, condition, notes FROM socket WHERE box=? COLLATE NOCASE", (a.box,))]
    if bases:
        print("\n  bases/sockets:")
        table(bases, ["base", "qty", "condition", "notes"])
    print()


def cmd_find(con, a):
    """Resolve a type name (and its equivalents) and print total stock plus a per-box breakdown for each match."""
    keys = resolve(con, a.type)
    if not keys:
        print(f"no type matching '{a.type}'")
        return
    for k in keys:
        t = con.execute("SELECT * FROM valve_type WHERE type_key=?", (k,)).fetchone()
        rows = [dict(r) for r in con.execute(
            "SELECT id, box, position, qty, manufacturer, condition, origin, notes "
            "FROM v_stock WHERE type_key=? ORDER BY CAST(box AS INTEGER), box, position", (k,))]
        total = sum(r["qty"] for r in rows)
        print(f"\n{t['name']}  -  {total} in stock across {len(rows)} box(es)")
        if t["function"]:
            print(f"  {t['function']}"
                  + (f", heater {t['heater_v']}V" if t["heater_v"] else "")
                  + (f", heater {t['heater_a']}A" if t["heater_a"] else ""))
        if t["equivalents"]:
            print(f"  equivalents: {t['equivalents']}")
        print()
        table(rows, shown(rows, ["id", "box", "position", "qty", "manufacturer",
                                 "condition", "origin"],
                          always={"id", "box", "qty", "manufacturer", "condition"}))
    print()


def cmd_show(con, a):
    """Print the full reference record for a type - parameters, equivalents, datasheet path, notes - plus its stock by box.

    If resolve() returns several matches (an ambiguous substring), only the
    first is shown in full; use a more specific name to disambiguate.
    """
    keys = resolve(con, a.type)
    if not keys:
        print(f"no type matching '{a.type}'")
        return
    t = con.execute("SELECT * FROM valve_type WHERE type_key=?", (keys[0],)).fetchone()
    print(f"\n{t['name']}   [{t['confidence']}]")
    print("=" * (len(t["name"]) + 16))
    fields = [("function", ""), ("family", ""), ("base", ""), ("pins", ""),
              ("heater_v", " V"), ("heater_a", " A"), ("va_max", " V"),
              ("pa_max", " W"), ("gm", " mA/V"), ("mu", ""),
              ("power_out", " W"), ("freq_max", " MHz")]
    for f, unit in fields:
        if t[f] is not None:
            print(f"  {f:<14} {t[f]}{unit}")
    if t["equivalents"]:
        print(f"  {'equivalents':<14} {t['equivalents']}")
    if t["datasheet_path"]:
        print(f"  {'datasheet':<14} {t['datasheet_path']}")
    if t["typical_use"]:
        print(f"\n  {t['typical_use']}")
    if t["notes"]:
        print(f"\n  notes: {t['notes'][:1500]}")
    rows = [dict(r) for r in con.execute(
        "SELECT id, box, position, qty, manufacturer, condition, origin "
        "FROM v_stock WHERE type_key=?", (keys[0],))]
    print(f"\n  stock ({sum(r['qty'] for r in rows)}):")
    table(rows, shown(rows, ["id", "box", "position", "qty", "manufacturer",
                             "condition", "origin"],
                      always={"id", "box", "qty", "manufacturer", "condition"}))
    print()


def cmd_search(con, a):
    """Filter stock lots by function/maker/box/free-text and numeric parameter comparisons, then print the matches.

    Builds the WHERE clause incrementally from whichever --options were
    given; "1=1" is a no-op base clause so the AND-join works even when no
    filters are supplied. Numeric filters (--heater, --pa, --va, --freq,
    --gm, --mu) are parsed by parse_cmp() so callers can pass a bare value
    (implicit '=') or a comparison like '>20'.
    """
    where, args = ["1=1"], []
    if a.function:
        where.append("(LOWER(t.function) LIKE ? OR LOWER(t.typical_use) LIKE ?)")
        args += [f"%{a.function.lower()}%"] * 2
    if a.maker:
        where.append("LOWER(s.manufacturer) LIKE ?")
        args.append(f"%{a.maker.lower()}%")
    if a.box:
        where.append("s.box = ? COLLATE NOCASE")
        args.append(a.box)
    if a.position:
        where.append("LOWER(s.position) LIKE ?")
        args.append(f"%{a.position.lower()}%")
    if a.origin:
        where.append("LOWER(s.origin) LIKE ?")
        args.append(f"%{a.origin.lower()}%")
    if a.alt:
        # the designation as marked on the valve, either secondary column
        where.append("(LOWER(s.type1) LIKE ? OR LOWER(s.type2) LIKE ?)")
        args += [f"%{a.alt.lower()}%"] * 2
    for field, expr in (("t.heater_v", a.heater), ("t.pa_max", a.pa),
                        ("t.va_max", a.va), ("t.freq_max", a.freq),
                        ("t.gm", a.gm), ("t.mu", a.mu)):
        if expr:
            op, val = parse_cmp(expr)
            where.append(f"{field} {op} ?")
            args.append(val)
    if a.text:
        # reference text plus everything recorded against the lot itself, so
        # "the one out of the Bush" or a number only printed on the glass is
        # findable without knowing which field it was written into
        where.append("(LOWER(t.name) LIKE ? OR LOWER(t.typical_use) LIKE ? "
                     "OR LOWER(t.notes) LIKE ? OR LOWER(t.equivalents) LIKE ? "
                     "OR LOWER(s.notes) LIKE ? OR LOWER(s.type1) LIKE ? "
                     "OR LOWER(s.type2) LIKE ? OR LOWER(s.origin) LIKE ? "
                     "OR LOWER(s.test_values) LIKE ? OR LOWER(s.other) LIKE ?)")
        args += [f"%{a.text.lower()}%"] * 10

    sql = f"""SELECT t.name AS type, s.box, s.position, s.qty, s.manufacturer,
                     s.condition, s.origin, t.function, t.heater_v, t.pa_max
              FROM stock s JOIN valve_type t ON s.type_key=t.type_key
              WHERE {' AND '.join(where)}
              ORDER BY t.name, CAST(s.box AS INTEGER)"""
    rows = [dict(r) for r in con.execute(sql, args)]
    print(f"\n{len(rows)} lots, {sum(r['qty'] for r in rows)} valves\n")
    table(rows, shown(rows, ["type", "box", "position", "qty", "manufacturer",
                             "condition", "origin", "function", "heater_v", "pa_max"],
                      always={"type", "box", "qty", "manufacturer", "function",
                              "heater_v", "pa_max"}))
    print()


def cmd_add(con, a):
    """Record a new stock lot for a type, auto-creating the valve_type record (via V.classify) if the type is unseen.

    New types are inserted with confidence='inferred' since their parameters
    come from the heuristic classifier rather than a datasheet; use 'set
    --confirm' later once real parameters are known. The box row is
    upserted (INSERT OR IGNORE) so an unfamiliar box still shows up in
    box listings even before it has a real location note. The per-lot detail
    fields (LOT_FIELDS - position, type1/type2, origin, test values, other)
    are all optional and can equally be filled in afterwards with 'edit'.
    """
    key = V.norm(a.type)
    exists = con.execute("SELECT 1 FROM valve_type WHERE type_key=?", (key,)).fetchone()
    if not exists:
        inf = V.classify(a.type)
        con.execute("""INSERT INTO valve_type (type_key,name,function,family,
                       heater_v,heater_a,confidence) VALUES (?,?,?,?,?,?,'inferred')""",
                    (key, a.type.strip(), inf.get("function"), inf.get("family"),
                     inf.get("heater_v"), inf.get("heater_a")))
        print(f"new type created: {a.type}"
              + (f"  ({inf.get('function')})" if inf.get("function") else ""))
    cur = con.execute(
        """INSERT INTO stock (type_key,box,position,qty,manufacturer,condition,
                              type1,type2,origin,test_values,other,date_added,notes)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (key, a.box, a.position, a.qty, a.maker, a.condition,
         a.type1, a.type2, a.origin, a.test, a.other,
         datetime.date.today().isoformat(), a.notes))
    con.execute("INSERT OR IGNORE INTO box (box, location) VALUES (?,?)", (a.box, "attic"))
    con.commit()
    print(f"added {a.qty} x {a.type} to box {a.box}"
          + (f", position {a.position}" if a.position else "")
          + f"  (lot {cur.lastrowid} - use 'edit {cur.lastrowid}' to change it)")


def cmd_edit(con, a):
    """Update one stock lot in place, by lot id - the counterpart to 'set' for the reference table.

    Lot ids are printed by 'box', 'find' and 'show', and by 'add' when it
    creates one. Only the options actually given are written, so editing the
    position leaves the origin alone; pass an empty string ('--origin ""')
    to clear a field that was filled in by mistake.
    """
    row = con.execute("SELECT * FROM v_stock WHERE id=?", (a.id,)).fetchone()
    if not row:
        print(f"no lot with id {a.id} - lot ids are shown by 'box', 'find' and 'show'")
        return
    fields, args = [], []
    for col, opt, _label in LOT_FIELDS:
        val = getattr(a, opt, None)
        if val is not None:
            fields.append(f"{col}=?")
            args.append(val or None)      # "" clears the field
    for col, val in (("box", a.box), ("qty", a.qty), ("manufacturer", a.maker),
                     ("condition", a.condition), ("notes", a.notes)):
        if val is not None:
            fields.append(f"{col}=?")
            args.append(val if col == "qty" else (val or None))
    if not fields:
        print("nothing to change - pass at least one option (--help lists them)")
        return
    args.append(a.id)
    con.execute(f"UPDATE stock SET {','.join(fields)} WHERE id=?", args)
    if a.box:
        con.execute("INSERT OR IGNORE INTO box (box, location) VALUES (?,?)", (a.box, "attic"))
    con.commit()
    r = con.execute("SELECT * FROM v_stock WHERE id=?", (a.id,)).fetchone()
    print(f"\nlot {a.id}: {r['qty']} x {r['type']} in box {r['box']}"
          + (f", position {r['position']}" if r["position"] else ""))
    for col, _opt, label in LOT_FIELDS:
        if col != "position" and r[col]:
            print(f"  {label:<12} {r[col]}")
    for col, label in (("manufacturer", "Maker"), ("condition", "Condition"),
                       ("notes", "Notes")):
        if r[col]:
            print(f"  {label:<12} {r[col]}")
    print()


def cmd_import_csv(con, a):
    """Bulk-add stock from a CSV (see upload_template.csv for the columns).

    Only `type` and `box` are required. Every other column, including the
    per-lot detail fields (LOT_FIELDS), is optional and may be left out of
    the file altogether - a CSV written for an older version still imports
    unchanged.
    """
    import csv
    added_types = added_lots = 0
    errors = []
    with open(a.file, encoding="utf-8-sig", newline="") as f:
        for i, row in enumerate(csv.DictReader(f), start=2):
            t = (row.get("type") or "").strip()
            box = (row.get("box") or "").strip()
            if not t or not box:
                errors.append(f"row {i}: type and box are required, skipped")
                continue
            key = V.norm(t)
            try:
                qty = int(row.get("qty") or 1)
            except ValueError:
                errors.append(f"row {i}: bad qty {row.get('qty')!r}, used 1")
                qty = 1
            maker = (row.get("maker") or "").strip() or None
            condition = (row.get("condition") or "").strip() or None
            notes = (row.get("notes") or "").strip() or None
            # "test" in the CSV matches the --test option name; the column it
            # lands in is test_values
            extra = [(row.get("test" if col == "test_values" else col) or "").strip() or None
                     for col, _opt, _label in LOT_FIELDS]
            if not con.execute("SELECT 1 FROM valve_type WHERE type_key=?", (key,)).fetchone():
                inf = V.classify(t)
                con.execute(
                    """INSERT INTO valve_type (type_key,name,function,family,
                       heater_v,heater_a,confidence) VALUES (?,?,?,?,?,?,'inferred')""",
                    (key, t, inf.get("function"), inf.get("family"),
                     inf.get("heater_v"), inf.get("heater_a")))
                added_types += 1
            con.execute(
                """INSERT INTO stock (type_key,box,qty,manufacturer,condition,date_added,notes,
                                      position,type1,type2,origin,test_values,other)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                [key, box, qty, maker, condition, datetime.date.today().isoformat(), notes]
                + extra)
            con.execute("INSERT OR IGNORE INTO box (box, location) VALUES (?,?)", (box, "attic"))
            added_lots += 1
    con.commit()
    print(f"imported {added_lots} lots ({added_types} new types)")
    for e in errors:
        print(" ", e)


def cmd_take(con, a):
    """Remove --qty valves of a type from stock, taking from the largest lot(s) first until satisfied or exhausted.

    Only resolve()'s first match is used, so an ambiguous substring takes
    from whichever type sorts first - pass an exact name if that matters.
    Largest-lot-first minimises the number of lots touched (and avoids
    leaving lots too small to be worth tracking) rather than draining lots
    in box or date order.
    """
    keys = resolve(con, a.type)
    if not keys:
        print("no such type")
        return
    rows = list(con.execute(
        "SELECT id,box,qty FROM stock WHERE type_key=?"
        + (" AND box=? COLLATE NOCASE" if a.box else "") + " ORDER BY qty DESC",
        (keys[0], a.box) if a.box else (keys[0],)))
    if not rows:
        print("none in stock")
        return
    left = a.qty
    for r in rows:
        if left <= 0:
            break
        take = min(left, r["qty"])
        if take == r["qty"]:
            con.execute("DELETE FROM stock WHERE id=?", (r["id"],))
        else:
            con.execute("UPDATE stock SET qty=qty-? WHERE id=?", (take, r["id"]))
        print(f"  took {take} from box {r['box']}")
        left -= take
    if left:
        print(f"  short by {left}")
    con.commit()


def cmd_move(con, a):
    """Reassign all stock lots of a type in one box (--frm) to another box (--to).

    Uses only resolve()'s first match; if that type has several lots in the
    source box (e.g. different manufacturers) they are all moved together -
    as is --position, if given, so use 'edit' to place one lot on its own.
    """
    keys = resolve(con, a.type)
    if a.position is None:
        con.execute("UPDATE stock SET box=? WHERE type_key=? AND box=? COLLATE NOCASE",
                    (a.to, keys[0], a.frm))
    else:
        # a position is only meaningful within its own box, so moving to a new
        # box either takes a new one or clears the old one rather than
        # carrying a stale reference across
        con.execute("UPDATE stock SET box=?, position=? WHERE type_key=? AND box=? COLLATE NOCASE",
                    (a.to, a.position or None, keys[0], a.frm))
    con.execute("INSERT OR IGNORE INTO box (box, location) VALUES (?,?)", (a.to, "attic"))
    con.commit()
    print(f"moved {keys[0]} from box {a.frm} to box {a.to}"
          + (f", position {a.position}" if a.position else ""))


# ---------------------------------------------------------------- bases/sockets

def cmd_bases(con, a):
    """List base/socket stock, optionally filtered to a base-name substring and/or a box."""
    where, args = ["1=1"], []
    if a.base:
        where.append("LOWER(base) LIKE ?")
        args.append(f"%{a.base.lower()}%")
    if a.box:
        where.append("box = ? COLLATE NOCASE")
        args.append(a.box)
    rows = [dict(r) for r in con.execute(
        f"SELECT base, box, qty, condition, notes FROM socket WHERE {' AND '.join(where)} "
        "ORDER BY base, CAST(box AS INTEGER), box", args)]
    total = sum(r["qty"] for r in rows)
    print(f"\n{len(rows)} lots, {total} sockets/bases\n")
    table(rows, ["base", "box", "qty", "condition", "notes"])
    print()


def cmd_sock_add(con, a):
    """Add a lot of bases/sockets to a box (the socket-table equivalent of cmd_add)."""
    con.execute("INSERT INTO socket (base,box,qty,condition,notes) VALUES (?,?,?,?,?)",
                (a.base.strip(), a.box, a.qty, a.condition, a.notes))
    con.execute("INSERT OR IGNORE INTO box (box, location) VALUES (?,?)", (a.box, "attic"))
    con.commit()
    print(f"added {a.qty} x {a.base} to box {a.box}")


def cmd_sock_take(con, a):
    """Remove qty bases/sockets of a given base type from stock, largest lot first (the socket-table equivalent of cmd_take)."""
    rows = list(con.execute(
        "SELECT id,box,qty FROM socket WHERE LOWER(base)=?"
        + (" AND box=? COLLATE NOCASE" if a.box else "") + " ORDER BY qty DESC",
        (a.base.lower(), a.box) if a.box else (a.base.lower(),)))
    if not rows:
        print("none in stock")
        return
    left = a.qty
    for r in rows:
        if left <= 0:
            break
        take = min(left, r["qty"])
        if take == r["qty"]:
            con.execute("DELETE FROM socket WHERE id=?", (r["id"],))
        else:
            con.execute("UPDATE socket SET qty=qty-? WHERE id=?", (take, r["id"]))
        print(f"  took {take} from box {r['box']}")
        left -= take
    if left:
        print(f"  short by {left}")
    con.commit()


def cmd_sock_move(con, a):
    """Reassign all base/socket lots of one base type in a box (--frm) to another box (--to) - the socket-table equivalent of cmd_move."""
    con.execute("UPDATE socket SET box=? WHERE LOWER(base)=? AND box=? COLLATE NOCASE",
                (a.to, a.base.lower(), a.frm))
    con.commit()
    print(f"moved {a.base} from box {a.frm} to box {a.to}")


def cmd_set(con, a):
    """Update a type's reference fields from whichever --options were supplied, leaving unset ones untouched.

    Only resolve()'s first match is edited. --confirm additionally sets
    confidence='confirmed', for recording that the parameters now come from
    a real datasheet rather than the heuristic classifier.
    """
    keys = resolve(con, a.type)
    if not keys:
        print("no such type")
        return
    fields, args = [], []
    for name in ("function", "base", "pins", "heater_v", "heater_a", "va_max",
                 "pa_max", "gm", "mu", "power_out", "freq_max", "typical_use",
                 "equivalents", "datasheet_path", "datasheet_url", "notes"):
        val = getattr(a, name, None)
        if val is not None:
            fields.append(f"{name}=?")
            args.append(val)
    if a.confirm:
        fields.append("confidence='confirmed'")
    if not fields:
        print("nothing to set")
        return
    args.append(keys[0])
    con.execute(f"UPDATE valve_type SET {','.join(fields)} WHERE type_key=?", args)
    con.commit()
    print(f"updated {keys[0]}")


def cmd_merge(con, a):
    """Fold one type into another, e.g. ECC83S -> ECC83, keeping all stock."""
    src, dst = V.norm(a.source), V.norm(a.into)
    s = con.execute("SELECT * FROM valve_type WHERE type_key=?", (src,)).fetchone()
    d = con.execute("SELECT * FROM valve_type WHERE type_key=?", (dst,)).fetchone()
    if not s or not d:
        print("both types must already exist")
        return
    n = con.execute("SELECT COALESCE(SUM(qty),0) c FROM stock WHERE type_key=?", (src,)).fetchone()["c"]
    if not a.yes:
        print(f"would move {n} valves from {s['name']} into {d['name']}, "
              f"recording '{s['name']}' as an equivalent. re-run with --yes")
        return
    eq = " ".join(sorted(set((d["equivalents"] or "").split()) | {s["name"]}))
    con.execute("UPDATE valve_type SET equivalents=? WHERE type_key=?", (eq, dst))
    con.execute("UPDATE stock SET type_key=?, notes=COALESCE(notes||' ','')||? "
                "WHERE type_key=?", (dst, f"[was listed as {s['name']}]", src))
    con.execute("DELETE FROM valve_type WHERE type_key=?", (src,))
    con.commit()
    print(f"merged {s['name']} into {d['name']} ({n} valves)")


def cmd_dupes(con, a):
    """Suggest types that look like variants of each other."""
    rows = list(con.execute("SELECT type_key, name FROM valve_type ORDER BY type_key"))
    keys = [r["type_key"] for r in rows]
    out = []
    for i, k in enumerate(keys):
        for k2 in keys[i + 1:]:
            if k2.startswith(k) and len(k2) - len(k) <= 2 and len(k) >= 3:
                out.append({"type": k, "variant": k2})
            elif k.startswith(k2) and len(k) - len(k2) <= 2 and len(k2) >= 3:
                out.append({"type": k2, "variant": k})
    print(f"\n{len(out)} possible duplicate pairs - review, then use 'merge':\n")
    table(out, ["type", "variant"])
    print()


def cmd_scan(con, a):
    """Walk the local datasheet archive and attach files to types."""
    root = a.archive
    if not os.path.isdir(root):
        print(f"archive directory not found: {root}")
        return
    index = {}
    for dirpath, _dirs, files in os.walk(root):
        for f in files:
            if not f.lower().endswith((".pdf", ".png", ".gif", ".jpg")):
                continue
            stem = V.norm(os.path.splitext(f)[0].split("~")[0])
            if stem:
                index.setdefault(stem, os.path.relpath(os.path.join(dirpath, f), root))
    n = 0
    for r in con.execute("SELECT type_key FROM valve_type WHERE datasheet_path IS NULL"):
        p = index.get(r["type_key"])
        if p:
            con.execute("UPDATE valve_type SET datasheet_path=? WHERE type_key=?",
                        (p, r["type_key"]))
            n += 1
    con.commit()
    print(f"archive files indexed: {len(index)}")
    print(f"types newly linked   : {n}")
    miss = con.execute("SELECT COUNT(*) c FROM valve_type WHERE datasheet_path IS NULL").fetchone()["c"]
    print(f"types still unlinked : {miss}")


def cmd_sheet(con, a):
    """Print the local path to a type's datasheet (joined with --archive) and, if --open was given, launch it with the OS's default viewer.

    If no datasheet is linked, prints links to two online datasheet
    archives instead. --open uses 'xdg-open' outside macOS, so it is a
    no-op on Windows.
    """
    keys = resolve(con, a.type)
    if not keys:
        print("no such type")
        return
    t = con.execute("SELECT name,datasheet_path FROM valve_type WHERE type_key=?", (keys[0],)).fetchone()
    if not t["datasheet_path"]:
        print(f"no local datasheet for {t['name']}")
        print(f"  try: https://frank.pocnet.net/sheets/  or  https://tdsl.duncanamps.com/show.php?des={t['name']}")
        return
    path = os.path.join(a.archive, t["datasheet_path"])
    print(path)
    if a.open:
        opener = "open" if sys.platform == "darwin" else "xdg-open"
        subprocess.run([opener, path], check=False)


def cmd_gaps(con, a):
    """Print two --limit-capped tables: in-stock types missing a datasheet, and in-stock types with no function classified - a to-do list for filling in reference data."""
    print("\ntypes held in stock with no datasheet linked:")
    rows = [dict(r) for r in con.execute("""
        SELECT t.name AS type, SUM(s.qty) qty, t.function
        FROM valve_type t JOIN stock s ON s.type_key=t.type_key
        WHERE t.datasheet_path IS NULL
        GROUP BY t.type_key ORDER BY qty DESC LIMIT ?""", (a.limit,))]
    table(rows, ["type", "qty", "function"])
    print("\ntypes with no function classified:")
    rows = [dict(r) for r in con.execute("""
        SELECT t.name AS type, SUM(s.qty) qty FROM valve_type t
        JOIN stock s ON s.type_key=t.type_key
        WHERE t.function IS NULL GROUP BY t.type_key ORDER BY qty DESC LIMIT ?""",
        (a.limit,))]
    table(rows, ["type", "qty"])
    print()


def cmd_docs(con, a):
    """List reference documents from the document table: the general
    library (type_key IS NULL) by default, or one type's primary datasheet
    plus its additional documents/links with --type."""
    if a.type:
        keys = resolve(con, a.type)
        if not keys:
            print(f"no type matching '{a.type}'")
            return
        t = con.execute("SELECT name, datasheet_path, datasheet_url FROM valve_type WHERE type_key=?",
                        (keys[0],)).fetchone()
        print(f"\n{t['name']}")
        if t["datasheet_path"]:
            print(f"  primary (local): {t['datasheet_path']}")
        elif t["datasheet_url"]:
            print(f"  primary (web):   {t['datasheet_url']}")
        else:
            print("  primary: none set")
        print()
        rows = [dict(r) for r in con.execute(
            "SELECT title, path, url, added FROM document WHERE type_key=? ORDER BY id", (keys[0],))]
        table(rows, ["title", "path", "url", "added"])
        print()
    else:
        rows = [dict(r) for r in con.execute(
            "SELECT title, path, url, added FROM document "
            "WHERE type_key IS NULL ORDER BY title")]
        print(f"\n{len(rows)} general reference document(s)\n")
        table(rows, ["title", "path", "url", "added"])
        print()


def cmd_stats(con, a):
    """Print a collection summary: headline counts, valves by function, and the fullest boxes."""
    q = lambda s: con.execute(s).fetchone()[0]
    print(f"\n  types            {q('SELECT COUNT(*) FROM valve_type')}")
    print(f"  stock lots       {q('SELECT COUNT(*) FROM stock')}")
    print(f"  valves total     {q('SELECT SUM(qty) FROM stock')}")
    print(f"  boxes in use     {q('SELECT COUNT(DISTINCT box) FROM stock')}")
    print(f"  datasheets held  {q('SELECT COUNT(*) FROM valve_type WHERE datasheet_path IS NOT NULL')}")
    print(f"""  confirmed params {q("SELECT COUNT(*) FROM valve_type WHERE confidence='confirmed'")}""")
    print(f"  bases/sockets    {q('SELECT COALESCE(SUM(qty),0) FROM socket')}")
    print("\n  by function:")
    rows = [dict(r) for r in con.execute("""
        SELECT COALESCE(t.function,'(unclassified)') AS function,
               COUNT(DISTINCT t.type_key) types, SUM(s.qty) valves
        FROM valve_type t JOIN stock s ON s.type_key=t.type_key
        GROUP BY 1 ORDER BY valves DESC LIMIT 20""")]
    table(rows, ["function", "types", "valves"])
    print("\n  fullest boxes:")
    rows = [dict(r) for r in con.execute("""
        SELECT box, COUNT(*) lots, SUM(qty) valves FROM stock
        GROUP BY box ORDER BY valves DESC LIMIT 10""")]
    table(rows, ["box", "lots", "valves"])
    print()


def cmd_export(con, a):
    """Write the full inventory to an .xlsx workbook with Stock, Types, and Other items sheets, formatted with bold headers, frozen header row, and auto-sized columns.

    Text fields are truncated to 300 characters per cell to keep the file
    manageable; column widths are sampled from the first 400 rows of each
    column rather than the whole sheet, for speed on large tables.
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Stock"
    cols = ["Box", "Position", "Type", "Type 1", "Type 2", "Qty", "Manufacturer",
            "Condition", "Origin", "Test values", "Other", "Function",
            "Heater V", "Pa max W", "Datasheet", "Notes"]
    ws.append(cols)
    for r in con.execute("""SELECT s.box, s.position, COALESCE(t.name,s.type_key),
                                   s.type1, s.type2, s.qty,
                                   s.manufacturer, s.condition, s.origin,
                                   s.test_values, s.other, t.function,
                                   t.heater_v, t.pa_max, t.datasheet_path, s.notes
                            FROM stock s LEFT JOIN valve_type t ON s.type_key=t.type_key
                            ORDER BY CAST(s.box AS INTEGER), s.box,
                                     s.position IS NULL, s.position, t.name"""):
        ws.append([(x[:300] if isinstance(x, str) else x) for x in r])

    ws2 = wb.create_sheet("Types")
    tcols = ["Type", "Function", "Family", "Base", "Heater V", "Heater A",
             "Va max", "Pa max W", "gm mA/V", "mu", "Power out W", "Freq max MHz",
             "Equivalents", "Datasheet", "Confidence", "In stock"]
    ws2.append(tcols)
    for r in con.execute("""SELECT t.name,t.function,t.family,t.base,t.heater_v,t.heater_a,
                                   t.va_max,t.pa_max,t.gm,t.mu,t.power_out,t.freq_max,
                                   t.equivalents,t.datasheet_path,t.confidence,
                                   (SELECT SUM(qty) FROM stock WHERE type_key=t.type_key)
                            FROM valve_type t ORDER BY t.name"""):
        ws2.append([(x[:300] if isinstance(x, str) else x) for x in r])

    ws3 = wb.create_sheet("Other items")
    ws3.append(["Box", "Description", "Qty", "Notes"])
    for r in con.execute("SELECT box, description, qty, notes FROM sundry"):
        ws3.append(list(r))

    for sh in wb.worksheets:
        for c in sh[1]:
            c.font = Font(name="Arial", bold=True)
            c.alignment = Alignment(horizontal="left")
        sh.freeze_panes = "A2"
        sh.auto_filter.ref = sh.dimensions
        for col in sh.columns:
            letter = col[0].column_letter
            width = max(len(str(c.value)) if c.value else 0 for c in col[:400])
            sh.column_dimensions[letter].width = min(max(width + 2, 9), 44)
        for row in sh.iter_rows(min_row=2):
            for c in row:
                c.font = Font(name="Arial")
    wb.save(a.path)
    print(f"written: {a.path}")


# ---------------------------------------------------------------- cli

LOT_OPTION_HELP = {
    "position": "where in the box, e.g. B-12 (row-column)",
    "type1": "secondary designation as marked, e.g. a US number",
    "type2": "a further secondary designation",
    "origin": "purchase, previous owner, or the set it came out of",
    "test": "what it measured on the tester",
    "other": "anything else: boxed/unboxed, printing, ...",
}


def add_lot_options(parser):
    """Declare the per-lot detail options (LOT_FIELDS) on `parser`.

    Shared by 'add' and 'edit' so the two always accept exactly the same
    set - the difference is only that 'add' writes them once and 'edit'
    writes whichever are given.
    """
    for _col, opt, _label in LOT_FIELDS:
        parser.add_argument("--" + opt, help=LOT_OPTION_HELP[opt])


def main():
    """Entry point: set up console output, build the argparse subcommand tree, then open the DB and dispatch to the chosen cmd_*() handler.

    Each subcommand's --help text and options are defined here; the actual
    work lives in the matching cmd_*() function, wired up via
    set_defaults(fn=...) and invoked as a.fn(con, a).
    """
    # Let output be piped into head/less without a BrokenPipeError traceback.
    try:
        from signal import signal, SIGPIPE, SIG_DFL
        signal(SIGPIPE, SIG_DFL)
    except (ImportError, ValueError):
        pass  # not available on Windows

    # Notes carry scraped text with non-Latin scripts (Cyrillic type names,
    # mu signs); Windows consoles default stdout to cp1252, which can't
    # encode them and crashes the print. UTF-8 with replacement is safe
    # everywhere and matches the DB's own encoding.
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except AttributeError:
        pass  # e.g. stdout replaced by something without reconfigure

    p = argparse.ArgumentParser(description="valve stock inventory")
    p.add_argument("--db", default=V.DB_DEFAULT)
    p.add_argument("--archive", default=V.ARCHIVE_DEFAULT)
    sub = p.add_subparsers(dest="cmd", required=True)

    s = sub.add_parser("box", help="list a box's contents");  s.add_argument("box");  s.set_defaults(fn=cmd_box)
    s = sub.add_parser("find", help="which boxes hold a type"); s.add_argument("type"); s.set_defaults(fn=cmd_find)
    s = sub.add_parser("show", help="full reference record");  s.add_argument("type"); s.set_defaults(fn=cmd_show)

    s = sub.add_parser("search", help="filter by parameters")
    s.add_argument("--function"); s.add_argument("--maker"); s.add_argument("--box")
    s.add_argument("--position", help="position within the box, e.g. B-12")
    s.add_argument("--origin", help="where it came from")
    s.add_argument("--alt", help="a secondary designation (Type 1 / Type 2)")
    s.add_argument("--heater", help="e.g. 6.3 or '<7'")
    s.add_argument("--pa", help="anode dissipation, e.g. '>20'")
    s.add_argument("--va"); s.add_argument("--freq"); s.add_argument("--gm"); s.add_argument("--mu")
    s.add_argument("--text", help="free text over name, use, notes, equivalents")
    s.set_defaults(fn=cmd_search)

    s = sub.add_parser("add", help="add stock")
    s.add_argument("type"); s.add_argument("--box", required=True)
    s.add_argument("--qty", type=int, default=1); s.add_argument("--maker")
    s.add_argument("--condition"); s.add_argument("--notes")
    add_lot_options(s)
    s.set_defaults(fn=cmd_add)

    s = sub.add_parser("edit", help="change one stock lot (by lot id, as shown by box/find/show)")
    s.add_argument("id", type=int, help="lot id, from the ID column of box/find/show")
    s.add_argument("--box"); s.add_argument("--qty", type=int); s.add_argument("--maker")
    s.add_argument("--condition"); s.add_argument("--notes")
    add_lot_options(s)
    s.set_defaults(fn=cmd_edit)

    s = sub.add_parser("import-csv", help="bulk-add stock from a CSV (see upload_template.csv)")
    s.add_argument("file"); s.set_defaults(fn=cmd_import_csv)

    s = sub.add_parser("take", help="remove stock you have used")
    s.add_argument("type"); s.add_argument("--qty", type=int, default=1)
    s.add_argument("--box"); s.set_defaults(fn=cmd_take)

    s = sub.add_parser("move", help="move a type between boxes")
    s.add_argument("type"); s.add_argument("--frm", required=True)
    s.add_argument("--to", required=True)
    s.add_argument("--position", help="position in the new box; '' clears the old one")
    s.set_defaults(fn=cmd_move)

    s = sub.add_parser("bases", help="list valve base/socket stock")
    s.add_argument("--base"); s.add_argument("--box"); s.set_defaults(fn=cmd_bases)

    s = sub.add_parser("docs", help="list reference documents/links (general library, or --type for one valve)")
    s.add_argument("--type", help="show one type's primary datasheet plus its extra documents/links")
    s.set_defaults(fn=cmd_docs)

    s = sub.add_parser("sock-add", help="add base/socket stock")
    s.add_argument("base"); s.add_argument("--box", required=True)
    s.add_argument("--qty", type=int, default=1)
    s.add_argument("--condition"); s.add_argument("--notes")
    s.set_defaults(fn=cmd_sock_add)

    s = sub.add_parser("sock-take", help="remove base/socket stock")
    s.add_argument("base"); s.add_argument("--qty", type=int, default=1)
    s.add_argument("--box"); s.set_defaults(fn=cmd_sock_take)

    s = sub.add_parser("sock-move", help="move base/socket stock between boxes")
    s.add_argument("base"); s.add_argument("--frm", required=True)
    s.add_argument("--to", required=True); s.set_defaults(fn=cmd_sock_move)

    s = sub.add_parser("set", help="edit a type's reference parameters")
    s.add_argument("type")
    for f in ("function", "base", "typical_use", "equivalents",
              "datasheet_path", "datasheet_url", "notes"):
        s.add_argument("--" + f.replace("_", "-"), dest=f)
    for f in ("heater_v", "heater_a", "va_max", "pa_max", "gm", "mu",
              "power_out", "freq_max"):
        s.add_argument("--" + f.replace("_", "-").replace("-max", ""), dest=f, type=float)
    s.add_argument("--pins", type=int)
    s.add_argument("--confirm", action="store_true", help="mark as read from a datasheet")
    s.set_defaults(fn=cmd_set)

    s = sub.add_parser("merge", help="fold one type into another")
    s.add_argument("source"); s.add_argument("into")
    s.add_argument("--yes", action="store_true"); s.set_defaults(fn=cmd_merge)
    s = sub.add_parser("dupes", help="list possible duplicate type entries"); s.set_defaults(fn=cmd_dupes)

    s = sub.add_parser("scan", help="link local archive files to types"); s.set_defaults(fn=cmd_scan)
    s = sub.add_parser("sheet", help="locate a datasheet")
    s.add_argument("type"); s.add_argument("--open", action="store_true"); s.set_defaults(fn=cmd_sheet)
    s = sub.add_parser("gaps", help="what still needs data")
    s.add_argument("--limit", type=int, default=25); s.set_defaults(fn=cmd_gaps)
    s = sub.add_parser("stats", help="collection summary"); s.set_defaults(fn=cmd_stats)
    s = sub.add_parser("export", help="write an xlsx snapshot")
    s.add_argument("path", nargs="?", default="valve_inventory.xlsx"); s.set_defaults(fn=cmd_export)

    a = p.parse_args()
    # Fresh clone or export: valves.db doesn't exist yet, but there's a
    # data/valves.sql snapshot to build it from. V.init_db() below would
    # happily open a brand-new, silently EMPTY database instead - warn rather
    # than let that pass with no explanation. Non-blocking (just a stderr
    # note, no prompt) so this stays safe to use in scripts/automation - see
    # valves_gui.py's main() for the interactive equivalent.
    if not os.path.exists(a.db) and os.path.exists(
            os.path.join(os.path.dirname(os.path.abspath(__file__)), "data", "valves.sql")):
        print(f"note: {a.db} doesn't exist yet, but data/valves.sql does - "
              f"run 'python3 snapshot.py --restore' to build it from that, "
              f"or this will start empty.", file=sys.stderr)
    con = V.init_db(a.db)
    a.fn(con, a)
    con.close()


if __name__ == "__main__":
    main()
